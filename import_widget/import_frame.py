#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pickle
import sqlite3
import numpy as np
import pandas as pd
from PyQt5.QtWidgets import (QFrame, QGroupBox, QLineEdit, QComboBox, QSpinBox, QTableView, QPushButton, QFileDialog,
                             QFormLayout, QHBoxLayout, QVBoxLayout)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from import_widget.import_model import ImportModel
from import_widget.import_delegate import TypeDelegate


class ImportFrame(QFrame):
    conn: sqlite3.Connection = None
    path: str = './'
    setting_file_path: str = None
    import_wb: Workbook = None
    import_ws: Worksheet = None
    model = ImportModel([])
    type_delegate = TypeDelegate()

    def __init__(self, *args):
        super(ImportFrame, self).__init__(*args)
        file_group = QGroupBox('Excel File')
        self.file_line = QLineEdit()
        self.file_line.setFixedWidth(750)
        self.file_line.setReadOnly(True)
        file_button = QPushButton('Load')
        file_button.setFixedWidth(100)
        file_layout = QHBoxLayout()
        file_layout.addWidget(self.file_line, alignment=Qt.AlignCenter)
        file_layout.addWidget(file_button, alignment=Qt.AlignCenter)
        file_group.setLayout(file_layout)
        file_group.setFixedWidth(884)

        import_group = QGroupBox('Import Settings')
        self.sheet_combo = QComboBox()
        self.header_spin = QSpinBox()
        self.header_spin.setRange(0, 0)
        self.name_line = QLineEdit()
        self.exists_combo = QComboBox()
        self.exists_combo.addItems(['append', 'replace', 'fail'])
        self.exists_combo.setStatusTip(
            '''if table also in database\nappend: insert new records\nreplace: replace data\nfail: abort import''')
        form_layout = QFormLayout()
        form_layout.addRow('Select Sheet', self.sheet_combo)
        form_layout.addRow('Select Title Row', self.header_spin)
        form_layout.addRow('Table Name', self.name_line)
        form_layout.addRow('If Table Exists', self.exists_combo)
        self.column_view = QTableView()
        self.column_view.setFixedSize(860, 400)
        save_button = QPushButton('&Save')
        save_button.setFixedWidth(100)
        load_button = QPushButton('&Load')
        load_button.setFixedWidth(100)
        fast_button = QPushButton('&FastLoad')
        fast_button.setFixedWidth(100)
        import_button = QPushButton('&Import')
        import_button.setFixedWidth(100)
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        button_layout.addWidget(save_button, alignment=Qt.AlignRight)
        button_layout.addWidget(load_button, alignment=Qt.AlignRight)
        button_layout.addWidget(fast_button, alignment=Qt.AlignRight)
        button_layout.addWidget(import_button, alignment=Qt.AlignRight)
        setting_layout = QVBoxLayout()
        setting_layout.addLayout(form_layout)
        setting_layout.addWidget(self.column_view, alignment=Qt.AlignCenter)
        setting_layout.addLayout(button_layout)
        import_group.setLayout(setting_layout)

        main_layout = QVBoxLayout()
        main_layout.addWidget(file_group, alignment=Qt.AlignCenter)
        main_layout.addWidget(import_group, alignment=Qt.AlignCenter)
        main_layout.addStretch()
        self.setLayout(main_layout)

        file_button.clicked.connect(self.load_file)
        self.header_spin.valueChanged.connect(self.load_column)
        save_button.clicked.connect(self.save_setting)
        load_button.clicked.connect(self.load_setting)
        fast_button.clicked.connect(self.auto_load)
        import_button.clicked.connect(self.import_data)

    def load_file(self):
        file_path: str = QFileDialog.getOpenFileName(self.parent(), 'Load File', self.path, 'Excel File(*.xlsx)',
                                                     options=QFileDialog.DontConfirmOverwrite)[0]
        if file_path:
            self.file_line.setText(file_path)
            self.path = '/'.join(file_path.split('/')[:-1])
            self.import_wb = load_workbook(file_path, data_only=True)
            self.sheet_combo.disconnect()
            self.sheet_combo.clear()
            self.sheet_combo.addItems(self.import_wb.sheetnames)
            self.load_sheet()
            self.sheet_combo.currentIndexChanged.connect(self.load_sheet)

    def load_sheet(self):
        self.import_ws = self.import_wb[self.sheet_combo.currentText()]
        self.name_line.setText(self.sheet_combo.currentText().replace(' ', '_', 10))
        self.header_spin.setRange(1, self.import_ws.max_row)
        self.load_column()

    def load_column(self):
        self.model = ImportModel([cell.value for cell in list(self.import_ws.rows)[self.header_spin.value() - 1]])
        self.column_view.setModel(self.model)
        self.column_view.setItemDelegateForColumn(1, self.type_delegate)
        self.column_view.setColumnWidth(0, 700)
        self.column_view.setColumnWidth(1, 125)

    def import_data(self):
        self.sender().setEnabled(False)
        df = pd.read_excel(self.file_line.text(), self.sheet_combo.currentText(), header=self.header_spin.value() - 1,
                           usecols=self.model.usecols, dtype=self.model.dtype, parse_dates=self.model.parse_dates,
                           date_parser=lambda x: pd.to_datetime(x).strftime('%Y-%m-%d'))
        for k, v in self.model.dtype.items():
            if v == str:
                df[k] = df[k].replace('nan', np.nan)
        df.to_sql(self.name_line.text(), self.conn, if_exists=self.exists_combo.currentText())
        self.sender().setEnabled(True)

    def save_setting(self):
        file_path: str = QFileDialog.getSaveFileName(self.parent(), 'Save Setting', './setting/', 'Setting File(*.stg)',
                                                     options=QFileDialog.DontConfirmOverwrite)[0]
        if file_path:
            save_settings = [self.name_line.text(), self.header_spin.value(), self.model.table_settings]
            file = open(file_path, 'wb')
            pickle.dump(save_settings, file)
            file.close()

    def load_setting(self):
        file_path: str = QFileDialog.getOpenFileName(self.parent(), 'Load Setting', './setting/', 'Setting File(*.stg)',
                                                     options=QFileDialog.DontConfirmOverwrite)[0]
        if file_path:
            self.setting_file_path = file_path
            file = open(file_path, 'rb')
            load_settings = pickle.load(file)
            self.name_line.setText(load_settings[0])
            self.header_spin.setValue(load_settings[1])
            self.model.table_settings = load_settings[2]
            self.column_view.reset()

    def auto_load(self):
        if self.setting_file_path:
            file = open(self.setting_file_path, 'rb')
            load_settings = pickle.load(file)
            self.name_line.setText(load_settings[0])
            self.header_spin.setValue(load_settings[1])
            self.model.table_settings = load_settings[2]
            self.column_view.reset()
